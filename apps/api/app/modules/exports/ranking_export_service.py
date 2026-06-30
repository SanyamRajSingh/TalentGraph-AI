from io import BytesIO
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.repositories.candidate_repository import CandidateRepository
from app.repositories.explanation_repository import ExplanationRepository
from app.repositories.ranking_repository import RankingRepository
from app.repositories.evaluation_repository import EvaluationRepository

class RankingExportService:
    """Builds comprehensive XLSX exports from existing data."""

    def __init__(
        self,
        ranking_repository: RankingRepository,
        candidate_repository: CandidateRepository,
        explanation_repository: ExplanationRepository,
        evaluation_repository: EvaluationRepository = None,
    ) -> None:
        self.ranking_repository = ranking_repository
        self.candidate_repository = candidate_repository
        self.explanation_repository = explanation_repository
        self.evaluation_repository = evaluation_repository

    def export_rankings(self, role_id: str, persona: str | None = None) -> bytes:
        rankings = self.ranking_repository.list_by_role_id(role_id, persona=persona)
        
        workbook = Workbook()
        
        # 1. Executive Summary
        ws_exec = workbook.active
        ws_exec.title = "Executive Summary"
        self._build_executive_summary(ws_exec, role_id, rankings)

        # 2. Rankings
        ws_rank = workbook.create_sheet(title="Rankings")
        self._build_rankings(ws_rank, rankings)

        # 3. Evaluation Breakdown
        ws_eval = workbook.create_sheet(title="Evaluation Breakdown")
        self._build_evaluation_breakdown(ws_eval, rankings)

        # 4. Explanations
        ws_exp = workbook.create_sheet(title="Explanations")
        self._build_explanations(ws_exp, rankings)

        # 5. Candidate Profiles
        ws_prof = workbook.create_sheet(title="Candidate Profiles")
        self._build_candidate_profiles(ws_prof, rankings)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _apply_header_style(self, sheet: Worksheet, headers: list[str]):
        sheet.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Freeze top row
        sheet.freeze_panes = "A2"
        # Add auto filter
        sheet.auto_filter.ref = sheet.dimensions

    def _format_sheet(self, sheet: Worksheet):
        # Auto size
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        lines = str(cell.value).split("\n")
                        max_length = max(max_length, max(len(line) for line in lines))
                except:
                    pass
            sheet.column_dimensions[column].width = min(max_length + 2, 80)

        # Top align and wrap text for all cells
        for row in sheet.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                # Alternating row colors
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F8F9FA")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")

    def _color_score(self, cell, score):
        if not isinstance(score, (int, float)):
            return
        if score >= 80:
            cell.fill = PatternFill("solid", fgColor="D4EDDA") # Green
            cell.font = Font(color="155724")
        elif score >= 50:
            cell.fill = PatternFill("solid", fgColor="FFF3CD") # Yellow
            cell.font = Font(color="856404")
        else:
            cell.fill = PatternFill("solid", fgColor="F8D7DA") # Red
            cell.font = Font(color="721C24")

    def _build_executive_summary(self, sheet, role_id, rankings):
        sheet.append(["TalentGraph AI - Executive Summary"])
        sheet.cell(1,1).font = Font(bold=True, size=16)
        sheet.append([])
        sheet.append(["Role ID", role_id])
        sheet.append(["Total Candidates Evaluated", len(rankings)])
        sheet.append([])
        
        headers = ["Top Candidates", "Match Score", "Recommendation"]
        self._apply_header_style(sheet, headers)
        
        for r in rankings[:10]:
            cand = self.candidate_repository.get_by_candidate_id(r.candidate_id)
            name = cand.name if cand else r.candidate_id
            rec = "Strong Hire" if r.confidence >= 75 else "Hire" if r.confidence >= 55 else "Growth Hire" if r.confidence >= 40 else "Borderline"
            sheet.append([name, r.score, rec])
            self._color_score(sheet.cell(sheet.max_row, 2), r.score)

        self._format_sheet(sheet)

    def _build_rankings(self, sheet, rankings):
        headers = ["Rank", "Candidate", "Match Score", "Confidence", "Recommendation", "Risk Profile", "Growth Stage"]
        self._apply_header_style(sheet, headers)
        
        for r in rankings:
            cand = self.candidate_repository.get_by_candidate_id(r.candidate_id)
            name = cand.name if cand else r.candidate_id
            rec = "Strong Hire" if r.confidence >= 75 else "Hire" if r.confidence >= 55 else "Growth Hire" if r.confidence >= 40 else "Borderline"
            stage = cand.growth_stage if cand else "Unknown"
            risk = "Low" if r.confidence >= 75 else "Medium" if r.confidence >= 50 else "High"
            
            sheet.append([r.rank, name, r.score, r.confidence, rec, risk, stage])
            self._color_score(sheet.cell(sheet.max_row, 3), r.score)
            self._color_score(sheet.cell(sheet.max_row, 4), r.confidence)

        self._format_sheet(sheet)

    def _build_evaluation_breakdown(self, sheet, rankings):
        headers = ["Candidate", "Technical Depth", "Learning Velocity", "Leadership", "Ownership", "Communication", "Project Complexity", "Collaboration"]
        self._apply_header_style(sheet, headers)
        
        for r in rankings:
            cand = self.candidate_repository.get_by_candidate_id(r.candidate_id)
            if cand:
                name = cand.name
                row = [name, cand.technical_depth, cand.learning_velocity, cand.leadership, cand.ownership, cand.communication, cand.project_complexity, cand.collaboration]
                sheet.append(row)
                for c in range(2, 9):
                    self._color_score(sheet.cell(sheet.max_row, c), row[c-1])
            else:
                sheet.append([r.candidate_id, "", "", "", "", "", "", ""])

        self._format_sheet(sheet)

    def _build_explanations(self, sheet, rankings):
        headers = ["Candidate", "Strengths", "Weaknesses", "Explanations", "Counterfactual Improvements"]
        self._apply_header_style(sheet, headers)
        
        for r in rankings:
            cand = self.candidate_repository.get_by_candidate_id(r.candidate_id)
            name = cand.name if cand else r.candidate_id
            exp = self.explanation_repository.get_by_candidate_id(r.candidate_id)
            if exp:
                sheet.append([
                    name,
                    "\n".join(exp.strengths),
                    "\n".join(exp.risks),
                    "\n".join(exp.reasoning),
                    "\n".join(exp.counterfactuals)
                ])
            else:
                eval_bundle = self.evaluation_repository.get(r.evaluation_id) if self.evaluation_repository and hasattr(r, "evaluation_id") else None
                if eval_bundle:
                    strengths = "\n".join(eval_bundle.technical.strengths + eval_bundle.growth.strengths)
                    risks = "\n".join(eval_bundle.technical.risks + eval_bundle.growth.risks)
                    sheet.append([name, strengths, risks, "(Detailed explanation not yet generated)", "(Counterfactuals not yet generated)"])
                else:
                    sheet.append([name, "", "", "", ""])

        self._format_sheet(sheet)

    def _build_candidate_profiles(self, sheet, rankings):
        headers = ["Candidate", "Email", "Phone", "Location", "Experience (Events)", "Domains", "Skills", "Career Trajectory"]
        self._apply_header_style(sheet, headers)
        
        all_candidates = self.candidate_repository.list_candidates()
        
        for cand in all_candidates:
            events = len(cand.timeline)
            timeline_str = "\n".join(f"{t.year}: {t.event}" for t in sorted(cand.timeline, key=lambda x: x.year, reverse=True))
            sheet.append([
                cand.name,
                cand.email or "",
                cand.phone or "",
                cand.location or "",
                f"{events} timeline events",
                ", ".join(cand.domains) if cand.domains else "",
                ", ".join(cand.skills) if cand.skills else "",
                timeline_str
            ])

        self._format_sheet(sheet)
