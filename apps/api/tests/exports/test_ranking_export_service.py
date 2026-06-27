from io import BytesIO

from openpyxl import load_workbook

from app.domain.candidate_twin import CandidateDigitalTwin
from app.domain.explanation import ExplanationProfile
from app.domain.ranking import HiringPersona, RankingResult
from app.modules.exports import RankingExportService
from app.repositories.memory import (
    InMemoryCandidateRepository,
    InMemoryEvaluationRepository,
    InMemoryExplanationRepository,
    InMemoryRankingRepository,
)


def _make_service():
    candidate_repository = InMemoryCandidateRepository()
    ranking_repository = InMemoryRankingRepository()
    explanation_repository = InMemoryExplanationRepository()
    evaluation_repository = InMemoryEvaluationRepository()

    candidate_repository.save(CandidateDigitalTwin(candidate_id="candidate_export", name="Ada Export"))
    ranking_repository.save_many(
        [
            RankingResult(
                candidate_id="candidate_export",
                role_id="role_export",
                rank=1,
                persona=HiringPersona.STARTUP_FOUNDER,
                score=88,
                confidence=81,
                evaluation_id="evaluation_export",
            )
        ]
    )
    explanation_repository.save(
        ExplanationProfile(
            candidate_id="candidate_export",
            role_id="role_export",
            ranking_position=1,
            strengths=["Strong Python alignment."],
            risks=["Limited fintech evidence."],
        )
    )
    return RankingExportService(
        ranking_repository=ranking_repository,
        candidate_repository=candidate_repository,
        explanation_repository=explanation_repository,
        evaluation_repository=evaluation_repository,
    )


def test_ranking_export_service_generates_xlsx_rows() -> None:
    svc = _make_service()
    content = svc.export_rankings(role_id="role_export", persona="startup_founder")

    workbook = load_workbook(BytesIO(content))
    
    # Verify expected sheets exist
    assert "Rankings" in workbook.sheetnames
    assert "Executive Summary" in workbook.sheetnames
    assert "Evaluation Breakdown" in workbook.sheetnames
    assert "Explanations" in workbook.sheetnames
    assert "Candidate Profiles" in workbook.sheetnames

    # Check Rankings sheet headers
    sheet = workbook["Rankings"]
    headers = [sheet.cell(row=1, column=col).value for col in range(1, 8)]
    assert headers[0] == "Rank"
    assert headers[1] == "Candidate"
    assert headers[2] == "Match Score"

    # Check data row
    assert sheet["A2"].value == 1
    assert sheet["B2"].value == "Ada Export"
    assert sheet["C2"].value == 88
    assert sheet["D2"].value == 81

    # Check Explanations sheet
    exp_sheet = workbook["Explanations"]
    assert exp_sheet["A2"].value == "Ada Export"
    assert "Strong Python alignment." in (exp_sheet["B2"].value or "")
    assert "Limited fintech evidence." in (exp_sheet["C2"].value or "")
